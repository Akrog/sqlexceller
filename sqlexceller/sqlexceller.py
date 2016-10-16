# -*- coding: utf-8 -*-

from __future__ import print_function
import argparse
from datetime import datetime
import os
import six
import sys

import openpyxl as excel
import sqlalchemy

import sqlexceller


__version__ = '0.1.0'

DEFAULT_OUTPUT_FILE = 'output.xlsx'
DATE_FORMAT = '%Y-%m-%d'
DEFAULT_URL = None


class Config(object):
    class MyParser(argparse.ArgumentParser):
        """Class to print verbose help on error."""
        def error(self, message):
            self.print_help()
            print('\nerror %s\n' % message, file=sys.stderr)
            sys.exit(1)

    class KeyValueAction(argparse.Action):
        def __call__(self, parser, namespace, values, option_string=None):
            try:
                k, v = values.split("=", 1)
            except ValueError:
                raise argparse.ArgumentError(self, "Format must be key=value")

            getattr(namespace, self.dest)[k] = v

    def __init__(self):
        general_description = """sqlexceller tool
    The tool will execute SQL queries and generate an Excel file with the
    results.
"""

        general_epilog = """

Available dialects and drivers are:
  - postgresql:
    - psycopg2
    - pg8000
  - mysql:
    - mysqldb
    - mysqlconnector
    - oursql
  - oracle:
    - cx_oracle
  - mssql:
    - pyodbc
    - pymssql
  - sqlite

There are some default parameters that will always be present:
  - NUM_QUERY
  - QUERY_NAME
  - DATE
  - DAY
  - MONTH
  - YEAR

Usage examples:
  - Execute a simple query on a SQLite DB.
      sqlexceller query.sql -d sqlite:///example.db

  - Execute multiple queries with 2 different parameters and a custom output
    file on a PostgreSQL DB:
      Contents of query1.sql:
        SELECT *
        FROM stocks
        where transaction = :transaction;

      Contents of query2.sql:
        SELECT *
        FROM stocks
        where transaction = :transaction and product = :product;

      sqlexceller query1.sql query2.sql -p transaction=BUY -p product=HAT \\
          -o "report :trans (:MONTH-:DAY).xlsx"

      Generated file will be something like: "report BUY (10-16).xlsx"
"""
        parser = Config.MyParser(
            description=general_description,
            epilog=general_epilog, argument_default='',
            formatter_class=argparse.RawTextHelpFormatter,
            prog='sqlexceller')
        parser.add_argument('-v', '--version', help="show program's version "
                            "number and exit", action='version',
                            version=('SQLExceller v%s' %
                                     sqlexceller.__version__)),
        parser.add_argument('query_files', nargs='+', metavar='file',
                            help='SQL Query file')
        parser.add_argument('--output', '-o',  dest='output',
                            help='Output file', default=DEFAULT_OUTPUT_FILE)
        parser.add_argument('--db_connection_info', '-d', help='DB connection '
                            'information as an URL in the form of\n'
                            '  dialect[+driver]://username:password@host:port/'
                            'database.', default=DEFAULT_URL, dest='url')
        parser.add_argument('--param', '-p',  dest='params',
                            action=Config.KeyValueAction,
                            default={},
                            help=('Adds a parameter for the SQL queries. '
                                  'Parameter must\nbe specified as a key=value'
                                  ' pair.  This argument can\nbe repeated as '
                                  'many times as necessary.'))
        self._parser = parser
        self._config = parser.parse_args()
        if not self._valid_config():
            sys.exit(2)
        vars(self).update(vars(self._config))

    def _valid_config(self):
        if not self._config.url:
            self._parser.print_help()
            print('\nError, database connection information missing, please '
                  'provide it via --db_connection_info.', file=sys.stderr)
            return False
        return True


class File(object):
    def __init__(self, filename, params, *args, **kwargs):
        self.filename = filename

        name, ext = os.path.splitext(filename)
        self.name = os.path.basename(name)

        date = datetime.now()
        self.params = params.copy()
        params.update(YEAR=six.text_type(date.year),
                      MONTH=six.text_type(date.month),
                      DAY=six.text_type(date.day),
                      DATE=date.strftime(DATE_FORMAT))


class XlsFile(File):
    def __init__(self, filename, params, column_titles=True, start_row=1,
                 start_column=1, *args, **kwargs):
        # Transform filename using parameters
        self._raw_filename = filename
        for key, value in params.items():
            filename = filename.replace(':' + key, value)
        self._workbook = excel.Workbook()
        # When we create a workbook, it always has one sheet, so we remove it
        self._workbook.remove_sheet(self._workbook.active)
        self._sheets_datarows = []
        self._num_sheets = 0
        self._start_row = start_row
        self._start_column = start_column
        self._column_titles = column_titles
        super(XlsFile, self).__init__(filename, params, *args, **kwargs)

    def __enter__(self):
        return self

    def get_sheet(self, title):
        if title not in self._workbook.sheetnames:
            self._num_sheets += 1
            self._workbook.create_sheet(title=title)
            self._sheets_datarows.append(0)
        return self._workbook[title]

    def __exit__(self, exc_type, exc_value, traceback):
        self._workbook.save(self.filename)

    def write_row(self, data, row, column=None, sheet=None):
        column = column or self._start_row
        sheet = sheet or self._workbook.worksheets[-1]
        for i, value in enumerate(data):
            sheet.cell(row=row, column=column + i).value = value

    def write_query(self, query):
        sheet = self.get_sheet(query.sqlfile.name)

        row = self._start_row
        if self._column_titles:
            row += 1

        for row_data in query.data:
            self.write_row(row_data.values(), row, sheet=sheet)
            row += 1

        if row_data and self._column_titles:
            self.write_row(row_data.keys(), self._start_row, sheet=sheet)
            row -= 1

        self._sheets_datarows[self._num_sheets - 1] = row - self._start_row


class SQLFile(File):
    def __init__(self, filename, params, i, *args, **kwargs):
        super(SQLFile, self).__init__(filename, params, *args, **kwargs)
        self.i = i
        self.params['QUERY_NAME'] = self.name
        self.params['NUM_QUERY'] = i + 1

        try:
            with open(filename) as f:
                self.data = f.read().strip()
        except Exception as e:
            print('\nError reading file %(filename)s:\n\t%(error)s' %
                  {'filename': filename, 'error': e}, file=sys.stderr)
            sys.exit(3)


class Query(object):
    def __init__(self, engine, sqlfile):
        self.engine = engine
        self.sqlfile = sqlfile
        self._data = None

    @property
    def data(self):
        if self._data is None:
            self._data = self.engine.execute(self.sqlfile.data,
                                             **self.sqlfile.params)
        return self._data


def process_queries(cfg):
    engine = sqlalchemy.create_engine(cfg.url)

    sqlfiles = [SQLFile(filename, cfg.params, i + 1)
                for i, filename in enumerate(cfg.query_files)]
    queries = [Query(engine, sqlfile) for sqlfile in sqlfiles]

    with XlsFile(cfg.output, cfg.params) as output:
        for query in queries:
            output.write_query(query)


def main():
    cfg = Config()
    process_queries(cfg)


if __name__ == "__main__":
    main()
